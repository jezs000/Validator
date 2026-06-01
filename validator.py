# validator.py

import logging
import re
from pathlib import Path

import pandas as pd
import yaml

from dateutil.parser import parse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from schwifty import IBAN


def validate_file(
    input_path: str,
    output_dir: str,
    config_path: str = "config.yaml"
):


    with open(config_path, "r", encoding="utf-8") as f:
        config = yaml.safe_load(f)

    required_columns = config["required_columns"]
    valid_currencies = config["valid_currencies"]
    salesforce_mapping = config["salesforce_mapping"]

    output_dir = Path(output_dir)
    output_dir.mkdir(exist_ok=True)

    valid_csv = output_dir / "salesforce_import.csv"
    error_xlsx = output_dir / "invoice_errors.xlsx"



    def is_empty(value):
        return pd.isna(value) or str(value).strip() == ""

    def is_valid_date(value):
        try:
            parse(str(value), dayfirst=True)
            return True
        except Exception:
            return False

    def is_valid_amount(value):
        try:
            return float(value) > 0
        except Exception:
            return False

    def is_valid_currency(value):
            return (
                str(value).upper().strip()
                in valid_currencies
            )

    def is_valid_dic(value):

        if is_empty(value):
            return False

        value = str(value).strip().upper()

        return bool(
            re.fullmatch(
                r"CZ\d{8,10}",
                value
            )
        )

    def is_valid_iban(value):

        try:
            return IBAN(
                str(value).replace(" ", "")
            ).is_valid

        except Exception:
            return False

    def is_valid_ico(value):

        if is_empty(value):
            return False

        value = str(value).strip()

        if not re.fullmatch(r"\d{8}", value):
            return False

        digits = [int(x) for x in value]

        checksum = (
            digits[0] * 8 +
            digits[1] * 7 +
            digits[2] * 6 +
            digits[3] * 5 +
            digits[4] * 4 +
            digits[5] * 3 +
            digits[6] * 2
        )

        mod = checksum % 11

        if mod == 0:
            expected = 1
        elif mod == 1:
            expected = 0
        else:
            expected = 11 - mod

        return expected == digits[7]



    df = pd.read_excel(input_path)



    missing_columns = [
        c
        for c in required_columns
        if c not in df.columns
    ]

    if missing_columns:
        raise ValueError(
            f"Missing columns: {missing_columns}"
        )



    statuses = []
    errors = []

    for _, row in df.iterrows():

        row_errors = []


        for col in required_columns:

            if is_empty(row[col]):
                row_errors.append(
                    f"Missing: {col}"
                )


        if not is_valid_date(row["Issue Date"]):
            row_errors.append(
                "Invalid Issue Date"
            )

        if not is_valid_date(row["Due Date"]):
            row_errors.append(
                "Invalid Due Date"
            )


        if not is_valid_ico(
            row["Vendor Company ID"]
        ):
            row_errors.append(
                "Invalid ICO"
            )


        if (
            "Vendor VAT Number" in df.columns
            and
            not pd.isna(
                row["Vendor VAT Number"]
            )
        ):
            if not is_valid_dic(
                row["Vendor VAT Number"]
            ):
                row_errors.append(
                    "Invalid DIC"
                )

        if not is_valid_amount(
            row["Total Amount"]
        ):
            row_errors.append(
                "Invalid Amount"
            )


        if not is_valid_currency(
            row["Currency"]
        ):
            row_errors.append(
                "Invalid Currency"
            )



        if (
            "IBAN" in df.columns
            and
            not pd.isna(row["IBAN"])
        ):
            if not is_valid_iban(
                row["IBAN"]
            ):
                row_errors.append(
                    "Invalid IBAN"
                )

        statuses.append(
            "ERROR"
            if row_errors
            else "VALID"
        )

        errors.append(
            "; ".join(row_errors)
        )



    df["Validation Status"] = statuses
    df["Validation Errors"] = errors



    valid_df = df[
        df["Validation Status"] == "VALID"
    ].copy()

    error_df = df[
        df["Validation Status"] == "ERROR"
    ].copy()



    valid_df = valid_df.rename(
        columns=salesforce_mapping
    )

    valid_df.drop(
        columns=[
            "Validation Status",
            "Validation Errors"
        ],
        errors="ignore",
        inplace=True
    )

    valid_df.to_csv(
        valid_csv,
        index=False,
        encoding="utf-8-sig"
    )



    error_df.to_excel(
        error_xlsx,
        index=False
    )

    wb = load_workbook(error_xlsx)
    ws = wb.active

    red_fill = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    headers = [
        cell.value
        for cell in ws[1]
    ]

    status_idx = (
        headers.index(
            "Validation Status"
        ) + 1
    )

    for row in ws.iter_rows(min_row=2):

        if (
            row[
                status_idx - 1
            ].value
            == "ERROR"
        ):
            for cell in row:
                cell.fill = red_fill

    wb.save(error_xlsx)


    return {
        "total_rows": len(df),
        "valid_rows": len(valid_df),
        "error_rows": len(error_df),
        "valid_csv": str(valid_csv),
        "error_xlsx": str(error_xlsx),
        "errors_dataframe": error_df
    }